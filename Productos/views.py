from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Sum
from django.http import HttpResponse, FileResponse, JsonResponse
from django.db import models
from django.utils import timezone
from datetime import datetime
import io
import re
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from .models import Producto, Categoria, Proveedor, Lote, ProductoConSerial, Bodega, LogActividad, RecepcionLote, Ubicacion, Movimiento
from .forms import (
    ProductoForm, CategoriaForm, ProveedorForm, LoteCompletoForm, 
    RecepcionLoteForm, SerialForm, BodegaForm, UbicacionForm, MovimientoForm
)
from usuarios.decorators import admin_required, puede_editar_required


# ========== FUNCIÓN DE LOGS ==========

def registrar_log(usuario, accion, modelo, objeto_id, objeto_nombre, ip='', detalles=''):
    try:
        LogActividad.objects.create(
            usuario=usuario,
            accion=accion,
            modelo=modelo,
            objeto_id=objeto_id,
            objeto_nombre=objeto_nombre,
            detalles=detalles,
            ip=ip
        )
    except:
        pass


# ========== FUNCIÓN PARA REGISTRAR MOVIMIENTOS ==========
def registrar_movimiento(usuario, producto, tipo, cantidad, lote=None, descripcion='', ip=''):
    """Registra un movimiento en el historial"""
    try:
        Movimiento.objects.create(
            producto=producto,
            lote=lote,
            tipo=tipo,
            cantidad=cantidad,
            descripcion=descripcion,
            usuario=usuario,
            ip=ip
        )
    except Exception as e:
        print(f"Error al registrar movimiento: {e}")


# ========== FUNCIÓN FIFO PARA OBTENER LOTE MÁS ANTIGUO ==========
def obtener_lote_fifo(producto, cantidad_necesaria):
    """
    🔥 LÓGICA FIFO: Obtiene los lotes más antiguos disponibles para vender
    Retorna una lista de lotes con las cantidades a tomar de cada uno
    """
    lotes_disponibles = Lote.objects.filter(
        producto=producto,
        estado__in=['completado', 'parcial']
    ).exclude(
        estado='agotado'
    ).filter(
        cantidad_recibida__gt=models.F('cantidad_vendida')
    ).order_by('fecha_ingreso')
    
    lotes_a_usar = []
    cantidad_restante = cantidad_necesaria
    
    for lote in lotes_disponibles:
        disponible = lote.disponible
        if disponible > 0:
            if disponible >= cantidad_restante:
                lotes_a_usar.append({
                    'lote': lote,
                    'cantidad': cantidad_restante
                })
                cantidad_restante = 0
                break
            else:
                lotes_a_usar.append({
                    'lote': lote,
                    'cantidad': disponible
                })
                cantidad_restante -= disponible
    
    if cantidad_restante > 0:
        raise ValueError(f'No hay suficiente stock disponible. Faltan {cantidad_restante} unidades')
    
    return lotes_a_usar


# ========== VISTAS DE UBICACIONES ==========

@login_required
def listar_ubicaciones(request):
    """Lista de ubicaciones predefinidas"""
    ubicaciones = Ubicacion.objects.all().order_by('nombre')
    return render(request, 'productos/ubicaciones_lista.html', {'ubicaciones': ubicaciones})


@login_required
@puede_editar_required
def crear_ubicacion(request):
    """Crear nueva ubicación"""
    if request.method == 'POST':
        form = UbicacionForm(request.POST)
        if form.is_valid():
            ubicacion = form.save()
            registrar_log(request.user, 'crear', 'Ubicacion', ubicacion.id, ubicacion.nombre)
            messages.success(request, f'✅ Ubicación "{ubicacion.nombre}" creada exitosamente')
            return redirect('productos:ubicaciones')
    else:
        form = UbicacionForm()
    
    return render(request, 'productos/ubicaciones_form.html', {
        'form': form,
        'titulo': 'Nueva Ubicación'
    })


@login_required
@puede_editar_required
def editar_ubicacion(request, pk):
    """Editar ubicación"""
    ubicacion = get_object_or_404(Ubicacion, pk=pk)
    if request.method == 'POST':
        form = UbicacionForm(request.POST, instance=ubicacion)
        if form.is_valid():
            form.save()
            registrar_log(request.user, 'editar', 'Ubicacion', ubicacion.id, ubicacion.nombre)
            messages.success(request, f'✅ Ubicación "{ubicacion.nombre}" actualizada correctamente')
            return redirect('productos:ubicaciones')
    else:
        form = UbicacionForm(instance=ubicacion)
    
    return render(request, 'productos/ubicaciones_form.html', {
        'form': form,
        'titulo': 'Editar Ubicación',
        'ubicacion': ubicacion
    })


@login_required
@admin_required
def eliminar_ubicacion(request, pk):
    """Eliminar ubicación"""
    ubicacion = get_object_or_404(Ubicacion, pk=pk)
    if request.method == 'POST':
        nombre = ubicacion.nombre
        if ubicacion.productos.count() > 0:
            messages.error(request, f'❌ No se puede eliminar "{nombre}" porque tiene {ubicacion.productos.count()} productos asociados')
        else:
            ubicacion.delete()
            registrar_log(request.user, 'eliminar', 'Ubicacion', ubicacion.id, nombre)
            messages.success(request, f'✅ Ubicación "{nombre}" eliminada correctamente')
        return redirect('productos:ubicaciones')
    return render(request, 'productos/ubicaciones_eliminar.html', {'ubicacion': ubicacion})


# ========== VISTAS DE MOVIMIENTOS (HISTORIAL) ==========

@login_required
def historial_movimientos(request, producto_id=None):
    """Ver historial de movimientos de un producto o todos"""
    if producto_id:
        producto = get_object_or_404(Producto, pk=producto_id)
        movimientos = Movimiento.objects.filter(producto=producto).order_by('-fecha')
        titulo = f'Historial de movimientos - {producto.nombre}'
    else:
        producto = None
        movimientos = Movimiento.objects.all().order_by('-fecha')
        titulo = 'Historial General de Movimientos'
    
    # ====== APLICAR FILTROS ======
    tipo = request.GET.get('tipo')
    if tipo:
        movimientos = movimientos.filter(tipo=tipo)
    
    fecha_desde = request.GET.get('fecha_desde')
    if fecha_desde:
        movimientos = movimientos.filter(fecha__date__gte=fecha_desde)
    
    fecha_hasta = request.GET.get('fecha_hasta')
    if fecha_hasta:
        movimientos = movimientos.filter(fecha__date__lte=fecha_hasta)
    
    # ====== CALCULAR RESÚMENES (ANTES DE PAGINAR) ======
    resumen = {
        'total_entradas': movimientos.filter(tipo='entrada').aggregate(Sum('cantidad'))['cantidad__sum'] or 0,
        'total_salidas': movimientos.filter(tipo='salida').aggregate(Sum('cantidad'))['cantidad__sum'] or 0,
        'total_movimientos': movimientos.count(),
    }
    
    # ====== PAGINACIÓN ======
    paginator = Paginator(movimientos, 50)
    page = request.GET.get('page')
    movimientos_page = paginator.get_page(page)
    
    return render(request, 'productos/movimientos_historial.html', {
        'movimientos': movimientos_page,
        'producto': producto,
        'titulo': titulo,
        'resumen': resumen,
        'tipos': Movimiento.TIPO_CHOICES,
    })


# ========== VISTAS DE PRODUCTOS ==========

@login_required
def listar_productos(request):
    productos = Producto.objects.all().order_by('-id')
    categorias = Categoria.objects.all()
    ubicaciones = Ubicacion.objects.all()
    
    # 🔥 NUEVO: Obtener todas las marcas para el filtro
    marcas = Producto.objects.values_list('marca', flat=True).distinct().exclude(marca__isnull=True).exclude(marca='')
    
    busqueda = request.GET.get('buscar', '')
    if busqueda:
        productos = productos.filter(
            Q(nombre__icontains=busqueda) |
            Q(codigo__icontains=busqueda) |
            Q(marca__icontains=busqueda) |
            Q(codigo_barras__icontains=busqueda) |
            Q(unidades__serial__icontains=busqueda)  # 🔥 BÚSQUEDA POR SERIAL
        ).distinct()
    
    categoria_id = request.GET.get('categoria')
    if categoria_id:
        productos = productos.filter(categoria_id=categoria_id)
    
    ubicacion_id = request.GET.get('ubicacion')
    if ubicacion_id:
        productos = productos.filter(ubicacion_id=ubicacion_id)
    
    # 🔥 NUEVO: Filtro por marca
    marca_filter = request.GET.get('marca')
    if marca_filter:
        productos = productos.filter(marca=marca_filter)
    
    for producto in productos:
        producto.stock_calculado = producto.stock_total()
    
    context = {
        'productos': productos,
        'categorias': categorias,
        'ubicaciones': ubicaciones,
        'marcas': marcas,  # 🔥 NUEVO: Pasar marcas al template
        'busqueda': busqueda,
    }
    return render(request, 'productos/lista.html', context)


@login_required
@puede_editar_required
def crear_producto(request):
    if request.method == 'POST':
        form = ProductoForm(request.POST)
        if form.is_valid():
            producto = form.save(commit=False)
            producto.save()
            registrar_log(request.user, 'crear', 'Producto', producto.id, producto.nombre)
            messages.success(request, f'✅ Producto "{producto.nombre}" creado exitosamente')
            messages.info(request, 'ℹ️ Recuerda crear un lote para agregar stock al producto')
            return redirect('productos:listar')
    else:
        form = ProductoForm()
    
    categorias = Categoria.objects.all()
    ubicaciones = Ubicacion.objects.all()
    return render(request, 'productos/form.html', {
        'form': form,
        'titulo': 'Nuevo Producto',
        'categorias': categorias,
        'ubicaciones': ubicaciones,
        'es_creacion': True
    })


@login_required
@puede_editar_required
def editar_producto(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    
    if request.method == 'POST':
        form = ProductoForm(request.POST, instance=producto)
        if form.is_valid():
            producto = form.save()
            registrar_log(request.user, 'editar', 'Producto', producto.id, producto.nombre)
            messages.success(request, f'✅ Producto "{producto.nombre}" actualizado correctamente')
            return redirect('productos:listar')
    else:
        form = ProductoForm(instance=producto)
    
    categorias = Categoria.objects.all()
    ubicaciones = Ubicacion.objects.all()
    return render(request, 'productos/form.html', {
        'form': form,
        'titulo': 'Editar Producto',
        'producto': producto,
        'categorias': categorias,
        'ubicaciones': ubicaciones,
        'es_creacion': False
    })


@login_required
@admin_required
def eliminar_producto(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    if request.method == 'POST':
        nombre = producto.nombre
        registrar_log(request.user, 'eliminar', 'Producto', producto.id, nombre)
        producto.delete()
        messages.success(request, f'✅ Producto "{nombre}" eliminado correctamente')
        return redirect('productos:listar')
    return render(request, 'productos/eliminar.html', {'producto': producto})


# ========== CARGA MASIVA EXCEL ==========

@login_required
def cargar_productos_excel(request):
    if request.method == 'POST' and request.FILES.get('archivo'):
        archivo = request.FILES['archivo']
        
        if not archivo.name.endswith(('.xlsx', '.xls')):
            messages.error(request, '❌ Formato no válido. Use archivos .xlsx o .xls')
            return redirect('productos:carga_masiva')
        
        try:
            df = pd.read_excel(archivo)
            df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
            
            columnas_requeridas = ['codigo', 'nombre', 'precio_venta']
            columnas_faltantes = [col for col in columnas_requeridas if col not in df.columns]
            
            if columnas_faltantes:
                messages.error(request, f'❌ Faltan columnas: {", ".join(columnas_faltantes)}')
                return redirect('productos:carga_masiva')
            
            creados = 0
            errores = 0
            for idx, row in df.iterrows():
                try:
                    codigo = str(row['codigo']).strip()
                    nombre = str(row['nombre']).strip()
                    precio_venta = float(row['precio_venta'])
                    
                    if not Producto.objects.filter(codigo=codigo).exists():
                        producto = Producto(
                            codigo=codigo,
                            nombre=nombre,
                            precio_venta=precio_venta,
                            stock_minimo=5,
                            precio_compra=precio_venta * 0.7,
                        )
                        producto.save()
                        creados += 1
                    else:
                        errores += 1
                except Exception as e:
                    errores += 1
            
            messages.success(request, f'✅ {creados} productos importados correctamente. {errores} productos omitidos (duplicados o errores)')
        except Exception as e:
            messages.error(request, f'❌ Error al procesar el archivo: {str(e)}')
        
        return redirect('productos:listar')
    
    return render(request, 'productos/carga_masiva.html')


# ========== EXPORTAR EXCEL CON SELECCIÓN ==========

@login_required
def exportar_productos_excel(request):
    # ====== OBTENER PRODUCTOS SELECCIONADOS ======
    productos_ids = request.GET.getlist('productos')
    
    if productos_ids:
        productos = Producto.objects.filter(id__in=productos_ids).order_by('nombre')
        titulo = "Productos Seleccionados"
    else:
        productos = Producto.objects.all().order_by('nombre')
        titulo = "Todos los Productos"
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    headers = ["Código", "Código Barras", "Nombre", "Marca", "Modelo", "Categoría", 
               "Stock", "Stock Mínimo", "Precio Compra", "Precio Venta", "Ubicación"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    for row, producto in enumerate(productos, 2):
        ws.cell(row=row, column=1, value=producto.codigo)
        ws.cell(row=row, column=2, value=producto.codigo_barras or "")
        ws.cell(row=row, column=3, value=producto.nombre)
        ws.cell(row=row, column=4, value=producto.marca or "")
        ws.cell(row=row, column=5, value=producto.modelo or "")
        ws.cell(row=row, column=6, value=producto.categoria.nombre if producto.categoria else "")
        ws.cell(row=row, column=7, value=producto.stock_total())
        ws.cell(row=row, column=8, value=producto.stock_minimo)
        ws.cell(row=row, column=9, value=float(producto.precio_compra or 0))
        ws.cell(row=row, column=10, value=float(producto.precio_venta))
        ws.cell(row=row, column=11, value=producto.get_ubicacion_nombre())
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 20
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=productos_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response


# ========== REPORTE PDF CON SELECCIÓN ==========

@login_required
def reporte_productos_pdf(request):
    # ====== OBTENER PRODUCTOS SELECCIONADOS ======
    productos_ids = request.GET.getlist('productos')
    
    if productos_ids:
        productos = Producto.objects.filter(id__in=productos_ids).order_by('nombre')
        titulo_reporte = "Reporte de Productos Seleccionados"
    else:
        productos = Producto.objects.all().order_by('nombre')
        titulo_reporte = "Reporte de Inventario - SIGI"
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    
    styles = getSampleStyleSheet()
    titulo_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], alignment=TA_CENTER, spaceAfter=30)
    titulo = Paragraph(
        f"<b>{titulo_reporte}</b><br/><font size='10'>{datetime.now().strftime('%d/%m/%Y %H:%M')}</font>", 
        titulo_style
    )
    
    data = [['Código', 'Producto', 'Marca', 'Stock', 'Stock Mínimo', 'Precio Venta', 'Ubicación', 'Estado']]
    valor_total = 0
    productos_bajo_stock = 0
    
    for p in productos:
        stock = p.stock_total()
        estado_stock = "⚠️ Crítico" if stock <= p.stock_minimo else "✅ Normal"
        if stock <= p.stock_minimo:
            productos_bajo_stock += 1
            
        data.append([
            p.codigo, 
            p.nombre[:30] if p.nombre else '-', 
            p.marca[:20] if p.marca else '-', 
            str(stock),
            str(p.stock_minimo), 
            f"${p.precio_venta:,.2f}",
            p.get_ubicacion_nombre(),
            estado_stock
        ])
        valor_total += float(p.precio_venta or 0) * int(stock or 0)
    
    tabla = Table(data, repeatRows=1)
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#f8fafc'), colors.white]),
    ]))
    
    resumen = Paragraph(
        f"<b>Resumen:</b> "
        f"Total Productos: {productos.count()} | "
        f"Productos Críticos: {productos_bajo_stock} | "
        f"Valor Total Inventario: ${valor_total:,.2f}", 
        styles['Normal']
    )
    
    doc.build([titulo, Spacer(1, 20), tabla, Spacer(1, 20), resumen])
    
    buffer.seek(0)
    return FileResponse(
        buffer, 
        as_attachment=True, 
        filename=f'reporte_productos_{datetime.now().strftime("%Y%m%d")}.pdf'
    )


# ========== VISTA PARA SELECCIONAR PRODUCTOS PARA REPORTE ==========

@login_required
def seleccionar_productos_reporte(request):
    """Vista para seleccionar productos para exportar con filtros avanzados"""
    productos = Producto.objects.all().order_by('nombre')
    categorias = Categoria.objects.all()
    marcas = Producto.objects.values_list('marca', flat=True).distinct().exclude(marca__isnull=True).exclude(marca='')
    
    # 🔥 Filtros avanzados
    busqueda = request.GET.get('buscar', '')
    if busqueda:
        productos = productos.filter(
            Q(nombre__icontains=busqueda) |
            Q(codigo__icontains=busqueda) |
            Q(marca__icontains=busqueda)
        )
    
    categoria_id = request.GET.get('categoria')
    if categoria_id:
        productos = productos.filter(categoria_id=categoria_id)
    
    # 🔥 Filtro por marca
    marca_filter = request.GET.get('marca')
    if marca_filter:
        productos = productos.filter(marca=marca_filter)
    
    # 🔥 Filtro por serial
    serial_filter = request.GET.get('serial')
    if serial_filter:
        productos = productos.filter(unidades__serial__icontains=serial_filter).distinct()
    
    # 🔥 Filtro por rango de stock (convierte a lista para poder filtrar)
    stock_min = request.GET.get('stock_min')
    stock_max = request.GET.get('stock_max')
    estado_stock = request.GET.get('estado_stock')
    
    # Aplicar filtros de stock después de obtener los productos
    productos_list = list(productos)
    if stock_min:
        try:
            stock_min_int = int(stock_min)
            productos_list = [p for p in productos_list if p.stock_total() >= stock_min_int]
        except ValueError:
            pass
    
    if stock_max:
        try:
            stock_max_int = int(stock_max)
            productos_list = [p for p in productos_list if p.stock_total() <= stock_max_int]
        except ValueError:
            pass
    
    if estado_stock:
        if estado_stock == 'bajo':
            productos_list = [p for p in productos_list if p.stock_total() <= p.stock_minimo and p.stock_total() > 0]
        elif estado_stock == 'normal':
            productos_list = [p for p in productos_list if p.stock_total() > p.stock_minimo]
        elif estado_stock == 'sin':
            productos_list = [p for p in productos_list if p.stock_total() == 0]
    
    # Paginación
    paginator = Paginator(productos_list, 20)
    page = request.GET.get('page')
    productos_page = paginator.get_page(page)
    
    context = {
        'productos': productos_page,
        'categorias': categorias,
        'marcas': marcas,  # 🔥 NUEVO
        'busqueda': busqueda,
    }
    return render(request, 'productos/seleccionar_reportes.html', context)


# ========== VISTAS DE CATEGORÍAS ==========

@login_required
def listar_categorias(request):
    categorias = Categoria.objects.all().order_by('nombre')
    return render(request, 'productos/categorias_lista.html', {'categorias': categorias})


@login_required
def crear_categoria(request):
    if request.method == 'POST':
        form = CategoriaForm(request.POST)
        if form.is_valid():
            categoria = form.save()
            messages.success(request, f'✅ Categoría "{categoria.nombre}" creada exitosamente')
            return redirect('productos:categorias')
    else:
        form = CategoriaForm()
    return render(request, 'productos/categorias_form.html', {'form': form, 'titulo': 'Nueva Categoría'})


@login_required
def editar_categoria(request, pk):
    categoria = get_object_or_404(Categoria, pk=pk)
    if request.method == 'POST':
        form = CategoriaForm(request.POST, instance=categoria)
        if form.is_valid():
            form.save()
            messages.success(request, f'✅ Categoría "{categoria.nombre}" actualizada correctamente')
            return redirect('productos:categorias')
    else:
        form = CategoriaForm(instance=categoria)
    return render(request, 'productos/categorias_form.html', {'form': form, 'titulo': 'Editar Categoría', 'categoria': categoria})


@login_required
def eliminar_categoria(request, pk):
    categoria = get_object_or_404(Categoria, pk=pk)
    if request.method == 'POST':
        nombre = categoria.nombre
        if categoria.producto_set.count() > 0:
            messages.error(request, f'❌ No se puede eliminar "{nombre}" porque tiene {categoria.producto_set.count()} productos asociados')
        else:
            categoria.delete()
            messages.success(request, f'✅ Categoría "{nombre}" eliminada correctamente')
        return redirect('productos:categorias')
    return render(request, 'productos/categorias_eliminar.html', {'categoria': categoria})


# ========== VISTAS DE PROVEEDORES ==========

@login_required
def listar_proveedores(request):
    proveedores = Proveedor.objects.all().order_by('nombre')
    return render(request, 'productos/proveedores_lista.html', {'proveedores': proveedores})


@login_required
def crear_proveedor(request):
    if request.method == 'POST':
        form = ProveedorForm(request.POST)
        if form.is_valid():
            proveedor = form.save()
            messages.success(request, f'✅ Proveedor "{proveedor.nombre}" creado exitosamente')
            return redirect('productos:proveedores')
    else:
        form = ProveedorForm()
    return render(request, 'productos/proveedores_form.html', {'form': form, 'titulo': 'Nuevo Proveedor'})


@login_required
def editar_proveedor(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)
    if request.method == 'POST':
        form = ProveedorForm(request.POST, instance=proveedor)
        if form.is_valid():
            form.save()
            messages.success(request, f'✅ Proveedor "{proveedor.nombre}" actualizado correctamente')
            return redirect('productos:proveedores')
    else:
        form = ProveedorForm(instance=proveedor)
    return render(request, 'productos/proveedores_form.html', {'form': form, 'titulo': 'Editar Proveedor', 'proveedor': proveedor})


@login_required
def eliminar_proveedor(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)
    if request.method == 'POST':
        nombre = proveedor.nombre
        if proveedor.lotes.count() > 0:
            messages.error(request, f'❌ No se puede eliminar "{nombre}" porque tiene {proveedor.lotes.count()} lotes asociados')
        else:
            proveedor.delete()
            messages.success(request, f'✅ Proveedor "{nombre}" eliminada correctamente')
        return redirect('productos:proveedores')
    return render(request, 'productos/proveedores_eliminar.html', {'proveedor': proveedor})


# ========== VISTAS DE LOTES ==========

@login_required
def listar_lotes(request):
    """Lista de lotes"""
    lotes = Lote.objects.all().order_by('-created_at')
    
    estado = request.GET.get('estado')
    if estado:
        lotes = lotes.filter(estado=estado)
    
    paginator = Paginator(lotes, 20)
    page = request.GET.get('page')
    lotes = paginator.get_page(page)
    
    return render(request, 'productos/lotes_lista.html', {
        'lotes': lotes,
        'estado_actual': estado,
    })


@login_required
def crear_lote(request):
    """Crear nuevo lote con validaciones"""
    if request.method == 'POST':
        form = LoteCompletoForm(request.POST)
        if form.is_valid():
            lote = form.save(commit=False)
            lote.created_by = request.user
            lote.cantidad_recibida = 0
            lote.cantidad_vendida = 0
            lote.fecha_ingreso = timezone.now()
            lote.save()
            registrar_log(request.user, 'crear', 'Lote', lote.id, lote.codigo)
            messages.success(request, f'✅ Lote "{lote.codigo}" creado exitosamente')
            return redirect('productos:detalle_lote', pk=lote.id)
    else:
        form = LoteCompletoForm()
    
    return render(request, 'productos/lote_form.html', {
        'form': form,
        'titulo': 'Crear Lote'
    })


@login_required
def detalle_lote(request, pk):
    """Detalle del lote"""
    lote = get_object_or_404(Lote, pk=pk)
    seriales = lote.productos.all()
    
    resumen = {
        'total': seriales.count(),
        'disponible': seriales.filter(estado='disponible').count(),
        'vendido': seriales.filter(estado='vendido').count(),
        'danado': seriales.filter(estado='danado').count(),
        'devuelto': seriales.filter(estado='devuelto').count(),
        'reservado': seriales.filter(estado='reservado').count(),
    }
    
    return render(request, 'productos/lote_detalle.html', {
        'lote': lote,
        'seriales': seriales,
        'resumen': resumen,
        'porcentaje': lote.porcentaje_recibido,
    })


@login_required
def recibir_lote(request, pk):
    """Recibir productos del lote con seriales"""
    lote = get_object_or_404(Lote, pk=pk)
    
    if lote.estado == 'completado':
        messages.warning(request, '⚠️ Este lote ya está completado')
        return redirect('productos:detalle_lote', pk=lote.id)
    
    if request.method == 'POST':
        form = RecepcionLoteForm(request.POST)
        if form.is_valid():
            cantidad = form.cleaned_data['cantidad']
            seriales_text = form.cleaned_data['seriales']
            notas = form.cleaned_data['notas']
            
            seriales_lista = [s.strip() for s in seriales_text.strip().split('\n') if s.strip()]
            
            if len(seriales_lista) != cantidad:
                messages.error(request, f'❌ La cantidad de seriales ({len(seriales_lista)}) no coincide con la cantidad indicada ({cantidad})')
                return render(request, 'productos/lote_recibir.html', {'form': form, 'lote': lote})
            
            if lote.cantidad_recibida + cantidad > lote.cantidad_total:
                messages.error(request, f'❌ No puedes recibir más de {lote.restante} unidades. Restante: {lote.restante}')
                return render(request, 'productos/lote_recibir.html', {'form': form, 'lote': lote})
            
            if len(seriales_lista) != len(set(seriales_lista)):
                messages.error(request, '❌ Hay seriales duplicados en la lista')
                return render(request, 'productos/lote_recibir.html', {'form': form, 'lote': lote})
            
            existentes = ProductoConSerial.objects.filter(serial__in=seriales_lista)
            if existentes.exists():
                existentes_str = ', '.join(existentes.values_list('serial', flat=True)[:5])
                if existentes.count() > 5:
                    existentes_str += f' y {existentes.count() - 5} más'
                messages.error(request, f'❌ Los siguientes seriales ya existen: {existentes_str}')
                return render(request, 'productos/lote_recibir.html', {'form': form, 'lote': lote})
            
            import re
            for serial in seriales_lista:
                if not re.match(r'^[A-Za-z0-9\-_]{3,50}$', serial):
                    messages.error(request, f'❌ El serial "{serial}" tiene formato inválido (solo letras, números, guiones, mínimo 3 caracteres)')
                    return render(request, 'productos/lote_recibir.html', {'form': form, 'lote': lote})
            
            creados = 0
            for serial in seriales_lista:
                ProductoConSerial.objects.create(
                    serial=serial,
                    producto_base=lote.producto,
                    lote=lote,
                    estado='disponible',
                    notas=notas
                )
                creados += 1
            
            lote.cantidad_recibida += cantidad
            if lote.cantidad_recibida >= lote.cantidad_total:
                lote.estado = 'completado'
                lote.fecha_entrega = timezone.now().date()
            else:
                lote.estado = 'parcial'
            lote.save()
            
            RecepcionLote.objects.create(
                lote=lote,
                cantidad=cantidad,
                seriales=seriales_text,
                recibido_por=request.user,
                notas=notas
            )
            
            registrar_movimiento(
                usuario=request.user,
                producto=lote.producto,
                tipo='entrada',
                cantidad=cantidad,
                lote=lote,
                descripcion=f'Recepción de lote {lote.codigo}',
                ip=request.META.get('REMOTE_ADDR', '')
            )
            
            registrar_log(request.user, 'recibir', 'Lote', lote.id, f'{cantidad} unidades - {lote.codigo}')
            messages.success(request, f'✅ Lote recibido: {cantidad} productos con seriales creados correctamente')
            return redirect('productos:detalle_lote', pk=lote.id)
    else:
        form = RecepcionLoteForm()
    
    return render(request, 'productos/lote_recibir.html', {
        'form': form,
        'lote': lote
    })


# ========== VISTAS DE SERIALES ==========

@login_required
def listar_seriales(request):
    """Lista de seriales"""
    seriales = ProductoConSerial.objects.all().order_by('serial')
    
    estado = request.GET.get('estado')
    if estado:
        seriales = seriales.filter(estado=estado)
    
    producto_id = request.GET.get('producto')
    if producto_id:
        seriales = seriales.filter(producto_base_id=producto_id)
    
    busqueda = request.GET.get('q')
    if busqueda:
        seriales = seriales.filter(serial__icontains=busqueda)
    
    paginator = Paginator(seriales, 50)
    page = request.GET.get('page')
    seriales = paginator.get_page(page)
    
    productos = Producto.objects.all().order_by('nombre')
    
    return render(request, 'productos/seriales_lista.html', {
        'seriales': seriales,
        'productos': productos,
        'estado_actual': estado,
    })


@login_required
def editar_serial(request, pk):
    """Editar estado de un serial con validaciones"""
    serial = get_object_or_404(ProductoConSerial, pk=pk)
    
    if request.method == 'POST':
        form = SerialForm(request.POST, instance=serial)
        if form.is_valid():
            form.save()
            registrar_log(request.user, 'editar', 'Serial', serial.id, serial.serial)
            messages.success(request, f'✅ Serial "{serial.serial}" actualizado a {serial.get_estado_display()}')
            return redirect('productos:listar_seriales')
    else:
        form = SerialForm(instance=serial)
    
    return render(request, 'productos/serial_form.html', {
        'form': form,
        'serial': serial
    })


# ========== VISTAS DE BODEGAS ==========

@login_required
def listar_bodegas(request):
    bodegas = Bodega.objects.all().order_by('nombre')
    return render(request, 'productos/bodegas_lista.html', {'bodegas': bodegas})


@login_required
def crear_bodega(request):
    if request.method == 'POST':
        form = BodegaForm(request.POST)
        if form.is_valid():
            bodega = form.save()
            messages.success(request, f'✅ Bodega "{bodega.nombre}" creada exitosamente')
            return redirect('productos:bodegas')
    else:
        form = BodegaForm()
    return render(request, 'productos/bodegas_form.html', {'form': form, 'titulo': 'Nueva Bodega'})


@login_required
def editar_bodega(request, pk):
    bodega = get_object_or_404(Bodega, pk=pk)
    if request.method == 'POST':
        form = BodegaForm(request.POST, instance=bodega)
        if form.is_valid():
            form.save()
            messages.success(request, f'✅ Bodega "{bodega.nombre}" actualizada correctamente')
            return redirect('productos:bodegas')
    else:
        form = BodegaForm(instance=bodega)
    return render(request, 'productos/bodegas_form.html', {'form': form, 'titulo': 'Editar Bodega', 'bodega': bodega})


@login_required
def eliminar_bodega(request, pk):
    bodega = get_object_or_404(Bodega, pk=pk)
    if request.method == 'POST':
        nombre = bodega.nombre
        if bodega.productos.count() > 0:
            messages.error(request, f'❌ No se puede eliminar "{nombre}" porque tiene {bodega.productos.count()} productos asociados')
        else:
            bodega.delete()
            messages.success(request, f'✅ Bodega "{nombre}" eliminada correctamente')
        return redirect('productos:bodegas')
    return render(request, 'productos/bodegas_eliminar.html', {'bodega': bodega})


# ========== ESCÁNER CÓDIGO DE BARRAS ==========

@login_required
def buscar_por_codigo_barras(request):
    """Busca un producto por su código de barras (API para escáner)"""
    codigo = request.GET.get('codigo', '').strip()
    
    if not codigo:
        return JsonResponse({'error': 'No se proporcionó código'}, status=400)
    
    try:
        producto = Producto.objects.filter(
            Q(codigo_barras=codigo) | Q(codigo=codigo)
        ).first()
        
        if producto:
            return JsonResponse({
                'success': True,
                'id': producto.id,
                'codigo': producto.codigo,
                'codigo_barras': producto.codigo_barras,
                'nombre': producto.nombre,
                'marca': producto.marca,
                'modelo': producto.modelo,
                'stock_actual': producto.stock_total(),
                'precio_venta': float(producto.precio_venta),
                'categoria': producto.categoria.nombre if producto.categoria else '',
                'ubicacion': producto.get_ubicacion_nombre(),
            })
        else:
            return JsonResponse({'success': False, 'error': 'Producto no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ========== VENTA DE PRODUCTOS CON FIFO ==========

@login_required
@puede_editar_required
def vender_producto(request, pk):
    """Vender producto usando lógica FIFO (primero en entrar, primero en salir)"""
    producto = get_object_or_404(Producto, pk=pk)
    
    if request.method == 'POST':
        cantidad = int(request.POST.get('cantidad', 0))
        
        if cantidad <= 0:
            messages.error(request, '❌ La cantidad debe ser mayor a 0')
            return redirect('productos:listar')
        
        stock_total = producto.stock_total()
        if cantidad > stock_total:
            messages.error(request, f'❌ No hay suficiente stock. Disponible: {stock_total}')
            return redirect('productos:listar')
        
        try:
            lotes_a_usar = obtener_lote_fifo(producto, cantidad)
            
            for item in lotes_a_usar:
                lote = item['lote']
                cantidad_lote = item['cantidad']
                
                lote.cantidad_vendida += cantidad_lote
                lote.actualizar_estado()
                
                registrar_movimiento(
                    usuario=request.user,
                    producto=producto,
                    tipo='salida',
                    cantidad=cantidad_lote,
                    lote=lote,
                    descripcion=f'Venta FIFO - Lote {lote.codigo}',
                    ip=request.META.get('REMOTE_ADDR', '')
                )
                
                seriales_vendidos = ProductoConSerial.objects.filter(
                    lote=lote,
                    estado='disponible'
                )[:cantidad_lote]
                for serial in seriales_vendidos:
                    serial.estado = 'vendido'
                    serial.save()
            
            registrar_log(
                request.user, 
                'editar', 
                'Producto', 
                producto.id, 
                f'Venta de {cantidad} unidades (FIFO)'
            )
            messages.success(request, f'✅ Venta realizada: {cantidad} unidades de "{producto.nombre}" (FIFO)')
            
        except ValueError as e:
            messages.error(request, f'❌ Error en la venta: {str(e)}')
        except Exception as e:
            messages.error(request, f'❌ Error inesperado: {str(e)}')
        
        return redirect('productos:listar')
    
    return render(request, 'productos/vender_producto.html', {
        'producto': producto,
        'stock_actual': producto.stock_total()
    })